#!/usr/bin/env python3
"""Export tracker Excel rows to JSON for GitHub Pages dashboard."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser(description="Export tracker workbook to web JSON")
    parser.add_argument("--excel", default="data/mainboard_ipos.xlsx", help="Path to tracker Excel")
    parser.add_argument("--out", default="docs/mainboard_ipos_web.json", help="Output JSON path")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")

    headers = [str(c.value or "").strip() for c in ws[1]]
    if not headers:
        raise RuntimeError("Header row is missing")

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row: dict[str, Any] = {}
        empty = True
        for c, header in enumerate(headers, start=1):
            if not header:
                continue
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                empty = False
            row[header] = "" if val is None else str(val)

        if empty:
            continue

        row.setdefault("applied", "")
        row.setdefault("got_ipo", "")
        rows.append(row)

    payload = {
        "count": len(rows),
        "rows": rows,
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, indent=2)

    print(f"Exported rows: {len(rows)}")
    print(f"Saved: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Fetch and persist Mainboard IPO tracking data from free public sources."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

MONEYCONTROL_IPO_URL = "https://www.moneycontrol.com/ipo/"
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36"

CSV_COLUMNS = [
    # "company_code",
    "company_name",
    # "ipo_type", # Mainline for all rows, so not needed in output
    "ipo_status",
    "open_date",
    "close_date",
    "allotment_date",
    "listing_date",
    "lot_size",
    "price_band",
    "invested",
    "listing_price",
    "output",
    # "issue_size_rs",
    # "issue_size_cr",
    "total_subsc",
    # "qib_sub",
    # "nii_sub",
    # "retail_sub",
    # "last_price",
    "listing_gain",
    "listing_gain%",
    "source_url",
    "last_updated_ist",
]


STATUS_COLORS = {
    "Open": "FFFDE68A",
    "Upcoming": "FFBFDBFE",
    "Listed": "FFBBF7D0",
    "Closed": "FFE5E7EB",
}


def fetch_html(url: str) -> str:
    req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "text/html"})
    with urlopen(req, timeout=30) as resp:  # nosec B310 - URL is constant and trusted by script design
        return resp.read().decode("utf-8", errors="ignore")


def fetch_html_or_empty(url: str) -> str:
    try:
        return fetch_html(url)
    except Exception:
        return ""


def extract_next_data(html: str) -> dict[str, Any]:
    match = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', html, re.S)
    if not match:
        raise ValueError("Unable to find __NEXT_DATA__ payload on page")
    return json.loads(match.group(1))


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    num = parse_float(value)
    if num is None:
        return None
    return int(round(num))


def fmt_date(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for pattern in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, pattern).strftime("%b %d, %Y")
        except ValueError:
            continue
    return text


def fmt_price_band(item: dict[str, Any]) -> str:
    low = parse_float(item.get("from_issue_price"))
    high = parse_float(item.get("to_issue_price"))
    issue_price = parse_float(item.get("issue_price"))
    if low is not None and high is not None:
        if abs(low - high) < 1e-9:
            return str(int(round(low)))
        return f"{int(round(low))}-{int(round(high))}"
    if issue_price is not None:
        return str(int(round(issue_price)))
    issue_price_raw = item.get("issue_price")
    return str(issue_price_raw).strip() if issue_price_raw else ""


def normalize_status(raw_status: Any, bucket_name: str) -> str:
    text = str(raw_status or "").strip().lower()
    if text:
        return text.title()
    if bucket_name == "listedIpo":
        return "Listed"
    if bucket_name == "closedIpo":
        return "Closed"
    return "Open"


def source_url(item: dict[str, Any]) -> str:
    raw = str(item.get("url") or "").strip()
    if not raw:
        return MONEYCONTROL_IPO_URL
    if raw.startswith("http"):
        return raw
    if raw.startswith("/"):
        return f"https://www.moneycontrol.com/ipo{raw}"
    return f"https://www.moneycontrol.com/ipo/{raw}"


def fmt_optional_int(value: Any) -> str:
    number = parse_int(value)
    if number is None:
        return ""
    return str(number)


def extract_detail_page_fields(url: str) -> dict[str, str]:
    html = fetch_html_or_empty(url)
    if not html:
        return {}

    def cell_value(label: str) -> str:
        pat = rf"<tr><td[^>]*>\s*{re.escape(label)}\s*</td><td[^>]*>(.*?)</td></tr>"
        match = re.search(pat, html, re.I | re.S)
        if not match:
            return ""
        value = re.sub(r"<[^<]+?>", "", match.group(1))
        return value.strip()

    raw_issue_price = cell_value("Issue Price")
    raw_lot_size = cell_value("Lot Size")
    raw_allotment = cell_value("Allotment Date")

    issue_price = ""
    if raw_issue_price:
        cleaned = raw_issue_price.replace("₹", "").replace(",", "").strip()
        nums = re.findall(r"\d+(?:\.\d+)?", cleaned)
        if len(nums) >= 2:
            issue_price = f"{int(round(float(nums[0])))}-{int(round(float(nums[1])))}"
        elif len(nums) == 1:
            issue_price = str(int(round(float(nums[0]))))

    lot_size = ""
    if raw_lot_size:
        lot = parse_int(raw_lot_size)
        if lot is not None:
            lot_size = str(lot)

    return {
        "price_band": issue_price,
        "lot_size": lot_size,
        "allotment_date": fmt_date(raw_allotment),
    }


def build_ipodetail_url(company_name: str, source_url: str) -> str:
    if "-ipodetail" in source_url:
        return source_url

    code_match = re.search(r"/([A-Z0-9]{3,})/?$", source_url)
    if not code_match:
        return source_url

    code = code_match.group(1).lower()
    slug = re.sub(r"[^a-z0-9]+", "-", company_name.lower()).strip("-")
    if not slug:
        return source_url

    return f"https://www.moneycontrol.com/ipo/{slug}-{code}-ipodetail/"


def enrich_missing_from_detail_pages(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    for row in rows:
        needs_enrichment = not row.get("lot_size") or not row.get("price_band") or not row.get("allotment_date")
        if not needs_enrichment:
            continue
        src = row.get("source_url", "")
        if not src or "moneycontrol.com" not in src:
            continue
        detail_url = build_ipodetail_url(row.get("company_name", ""), src)
        detail = extract_detail_page_fields(detail_url)

        if not detail.get("lot_size") and detail_url != src:
            detail = extract_detail_page_fields(src)

        if detail.get("lot_size") and not row.get("lot_size"):
            row["lot_size"] = detail["lot_size"]
        if detail.get("price_band") and not row.get("price_band"):
            row["price_band"] = detail["price_band"]
        if detail.get("allotment_date") and not row.get("allotment_date"):
            row["allotment_date"] = detail["allotment_date"]
    return rows


def parse_price_band_high(price_band: str) -> float | None:
    text = str(price_band or "").strip()
    if not text:
        return None
    if "-" in text:
        parts = text.split("-")
        parsed = parse_int(parts[-1])
        return float(parsed) if parsed is not None else None
    parsed = parse_int(text)
    return float(parsed) if parsed is not None else None


def recalculate_derived_fields(row: dict[str, str]) -> dict[str, str]:
    lot_size = parse_int(row.get("lot_size"))
    price_high = parse_price_band_high(row.get("price_band", ""))
    listing_price = parse_int(row.get("listing_price"))

    invested = None
    if lot_size is not None and price_high is not None:
        invested = lot_size * price_high
        row["invested"] = str(int(round(invested)))
    else:
        row["invested"] = ""

    if lot_size is not None and listing_price is not None:
        output = lot_size * listing_price
        row["output"] = str(int(round(output)))
        if invested not in (None, 0):
            listing_gain = output - invested
            row["listing_gain"] = str(int(round(listing_gain)))
            row["listing_gain%"] = str(int(round((listing_gain / invested) * 100)))
    else:
        row["output"] = ""

    return row


def normalize_row(item: dict[str, Any], bucket_name: str, now_ist: str) -> dict[str, str]:
    total_sub = item.get("total_subs", item.get("total"))
    status = normalize_status(item.get("ipo_status"), bucket_name)
    listing_price = parse_int(item.get("dt_open")) if status == "Listed" else None

    return {
        # "company_code": str(item.get("company_code") or item.get("sc_id") or "").strip(),
        "company_name": str(item.get("company_name") or "").strip(),
        # "ipo_type": str(item.get("ipo_type") or "").strip(),
        "ipo_status": status,
        "open_date": fmt_date(item.get("open_date")),
        "close_date": fmt_date(item.get("close_date")),
        "allotment_date": fmt_date(item.get("allotment_date")),
        "listing_date": fmt_date(item.get("listing_date")),
        "lot_size": str(parse_int(item.get("lot_size")) or ""),
        "price_band": fmt_price_band(item),
        "invested": "",
        "listing_price": "" if listing_price is None else str(listing_price),
        "output": "",
        # "issue_size_rs": str(issue_size_rs or ""),
        # "issue_size_cr": issue_size_cr,
        "total_subsc": fmt_optional_int(total_sub),
        # "qib_sub": fmt_optional_int(item.get("qib")),
        # "nii_sub": fmt_optional_int(item.get("nii")),
        # "retail_sub": fmt_optional_int(item.get("retail")),
        # "last_price": fmt_optional_int(item.get("last_price")),
        "listing_gain": "",
        "listing_gain%": "",
        "source_url": source_url(item),
        "last_updated_ist": now_ist,
    }


def merge_rows(existing: dict[str, str], incoming: dict[str, str]) -> dict[str, str]:
    if existing.get("ipo_status", "") == "Listed":
        # Once a row is listed, keep it frozen.
        return dict(existing)

    merged = dict(existing)
    for key, value in incoming.items():
        if value not in ("", None):
            merged[key] = value

    status_rank = {"Upcoming": 1, "Open": 2, "Closed": 3, "Listed": 4}
    old_status = existing.get("ipo_status", "")
    new_status = incoming.get("ipo_status", "")
    if status_rank.get(new_status, 0) >= status_rank.get(old_status, 0):
        merged["ipo_status"] = new_status

    # Open/Upcoming/Closed rows should not carry listing-time fields.
    if merged.get("ipo_status", "") != "Listed":
        merged["listing_price"] = ""
        merged["output"] = ""
        merged["listing_gain"] = ""
        merged["listing_gain%"] = ""

    return merged


def row_key(row: dict[str, str]) -> str:
    return f"name:{row.get('company_name', '').strip().lower()}"


def load_existing_excel(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    wb = load_workbook(path)
    ws = wb.active
    if ws is None:
        return {}
    if ws.max_row < 2:
        return {}

    headers = [str(c.value or "").strip() for c in ws[1]]
    if not headers:
        return {}

    rows: dict[str, dict[str, str]] = {}
    for r in range(2, ws.max_row + 1):
        row: dict[str, str] = {}
        for i, header in enumerate(headers, start=1):
            if not header:
                continue
            value = ws.cell(row=r, column=i).value
            row[header] = "" if value is None else str(value)

        for col in CSV_COLUMNS:
            row.setdefault(col, "")

        key = row_key(row)
        if key:
            rows[key] = row

    return rows


def save_excel(path: Path, rows: list[dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws.title = "Mainboard IPOs"

    ws.append(CSV_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in CSV_COLUMNS])

    header_fill = PatternFill(fill_type="solid", fgColor="FF1F2937")
    header_font = Font(color="FFFFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    status_idx = CSV_COLUMNS.index("ipo_status") + 1
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(row=r, column=status_idx).value or "").strip()
        color = STATUS_COLORS.get(status)
        if color:
            ws.cell(row=r, column=status_idx).fill = PatternFill(fill_type="solid", fgColor=color)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 34,
        "B": 14,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 11,
        "H": 14,
        "I": 12,
        "J": 16,
        "K": 12,
        "L": 14,
        "M": 14,
        "N": 16,
        "O": 14,
        "P": 24,
        "Q": 10,
        "R": 10,
        "S": 14,
        "T": 12,
        "U": 12,
        "V": 50,
        "W": 21,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def collect_mainboard_rows(payload: dict[str, Any]) -> list[dict[str, str]]:
    ipo_data = payload["props"]["pageProps"]["ipoData"]
    buckets = ["open_Upcoming", "openIpoList", "closedIpo", "listedIpo"]

    ist = dt.timezone(dt.timedelta(hours=5, minutes=30))
    now_ist = dt.datetime.now(ist).strftime("%Y-%m-%d %H:%M:%S")
    collected: dict[str, dict[str, str]] = {}

    for bucket in buckets:
        items = ipo_data.get(bucket, [])
        if not isinstance(items, list):
            continue
        for item in items:
            if str(item.get("ipo_type", "")).strip().lower() != "mainline":
                continue
            row = normalize_row(item, bucket, now_ist)
            key = row_key(row)
            if not key:
                continue
            if key in collected:
                collected[key] = merge_rows(collected[key], row)
            else:
                collected[key] = row

    return list(collected.values())


def sort_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def parse_for_sort(value: str) -> dt.datetime:
        if not value:
            return dt.datetime(1900, 1, 1)
        try:
            return dt.datetime.strptime(value, "%b %d, %Y")
        except ValueError:
            return dt.datetime(1900, 1, 1)

    def row_sort_key(row: dict[str, str]) -> tuple[int, dt.datetime, str]:
        listed_dt = parse_for_sort(row.get("listing_date", ""))
        has_date = 0 if row.get("listing_date", "") else 1
        return (has_date, listed_dt, row.get("company_name", ""))

    return sorted(rows, key=row_sort_key)


def sanitize_dates(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    date_re = re.compile(r"^[A-Z][a-z]{2} \d{2}, \d{4}$")
    date_fields = ["open_date", "close_date", "allotment_date", "listing_date"]

    for row in rows:
        for field in date_fields:
            value = row.get(field, "").strip()
            if not value:
                row[field] = ""
                continue
            if not date_re.match(value):
                row[field] = ""
    return rows


def sanitize_numeric_columns(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    int_fields = [
        "lot_size",
        "invested",
        "listing_price",
        "output",
        "total_subsc",
        "listing_gain",
        "listing_gain%",
    ]

    for row in rows:
        price_band = row.get("price_band", "").strip()
        if price_band:
            if "-" in price_band:
                left, right = price_band.split("-", 1)
                lnum = parse_int(left)
                rnum = parse_int(right)
                if lnum is not None and rnum is not None:
                    row["price_band"] = f"{lnum}-{rnum}"
            else:
                num = parse_int(price_band)
                if num is not None:
                    row["price_band"] = str(num)

        for field in int_fields:
            value = row.get(field, "").strip()
            if not value:
                row[field] = ""
                continue
            num = parse_int(value)
            row[field] = "" if num is None else str(num)

    return rows


def persist_snapshot(snapshot_path: Path, payload: dict[str, Any]) -> None:
    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    with snapshot_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, indent=2)


def run(snapshot_json: Path, output_xlsx: Path) -> int:
    try:
        html = fetch_html(MONEYCONTROL_IPO_URL)
        payload = extract_next_data(html)
    except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError) as exc:
        print(f"ERROR: failed to fetch/parse source data: {exc}", file=sys.stderr)
        return 2

    incoming_rows = collect_mainboard_rows(payload)
    existing = load_existing_excel(output_xlsx)

    before_count = len(existing)
    for row in incoming_rows:
        key = row_key(row)
        if key in existing:
            existing[key] = merge_rows(existing[key], row)
        else:
            existing[key] = row

    visible_rows = [row for row in existing.values() if row.get("ipo_status", "") != "Closed"]

    merged_rows = sanitize_dates(sort_rows(visible_rows))
    merged_rows = enrich_missing_from_detail_pages(merged_rows)
    merged_rows = sort_rows(merged_rows)
    merged_rows = [recalculate_derived_fields(row) for row in merged_rows]
    merged_rows = sanitize_numeric_columns(merged_rows)
    save_excel(output_xlsx, merged_rows)
    persist_snapshot(snapshot_json, payload)

    after_count = len(merged_rows)
    new_count = max(0, after_count - before_count)
    print(f"Updated {output_xlsx}")
    print(f"Mainboard IPO rows: {after_count} (new this run: {new_count})")
    print(f"Snapshot saved: {snapshot_json}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Track Mainboard IPOs and update a local CSV")
    parser.add_argument(
        "--snapshot",
        default="data/source_snapshot.json",
        help="Path to raw source snapshot JSON (default: data/source_snapshot.json)",
    )
    parser.add_argument(
        "--excel",
        default="data/mainboard_ipos.xlsx",
        help="Path to output XLSX (default: data/mainboard_ipos.xlsx)",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    raise SystemExit(run(Path(args.snapshot), Path(args.excel)))

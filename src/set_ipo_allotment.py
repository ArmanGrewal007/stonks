#!/usr/bin/env python3
"""Set manual IPO allotment decision in the Excel tracker.

Values supported for got_ipo: yes, no, na, clear.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

import re

TARGET_COLUMN = "got_ipo"
SUPPORTED = {
    "yes": "Yes",
    "no": "No",
    "na": "N/A",
    "clear": "",
}


def normalize_company_name(name: str) -> str:
    text = str(name or "").strip().lower()
    text = re.sub(r"\b(private|pvt|limited|ltd|inc|corp|corporation|llp)\b", "", text)
    text = re.sub(r"[^a-z0-9]", "", text)
    return text


def find_col_index(headers: list[str], name: str) -> int:
    for idx, header in enumerate(headers, start=1):
        if header.strip() == name:
            return idx
    return -1


def ensure_column(ws, headers: list[str], col_name: str) -> int:
    idx = find_col_index(headers, col_name)
    if idx != -1:
        return idx

    # Insert after company_name when column is missing.
    company_idx = find_col_index(headers, "company_name")
    if company_idx == -1:
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = col_name
        return 1

    insert_at = company_idx + 1
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = col_name
    return insert_at


def main() -> int:
    parser = argparse.ArgumentParser(description="Set IPO applied / allotment status in Excel")
    parser.add_argument("--excel", default="data/mainboard_ipos.xlsx", help="Path to tracker Excel file")
    parser.add_argument("--company", required=True, help="Exact company_name to update")
    parser.add_argument(
        "--got-ipo",
        choices=sorted(SUPPORTED.keys()),
        help="Allotment decision: yes | no | na | clear",
    )
    parser.add_argument(
        "--applied",
        choices=sorted(SUPPORTED.keys()),
        help="Applied decision: yes | no | na | clear",
    )
    args = parser.parse_args()

    if not args.got_ipo and not args.applied:
        raise RuntimeError("Must specify at least one of --got-ipo or --applied")

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")

    headers = [str(c.value or "").strip() for c in ws[1]]
    if not headers:
        raise RuntimeError("Header row is missing")

    got_col = ensure_column(ws, headers, "got_ipo") if args.got_ipo else None
    headers = [str(c.value or "").strip() for c in ws[1]]
    applied_col = ensure_column(ws, headers, "applied") if args.applied else None
    headers = [str(c.value or "").strip() for c in ws[1]]

    company_col = find_col_index(headers, "company_name")
    if company_col == -1:
        raise RuntimeError("company_name column not found")

    wanted = args.company.strip().lower()
    wanted_norm = normalize_company_name(args.company)
    if not wanted:
        raise RuntimeError("--company cannot be empty")

    updated_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        company = str(ws.cell(row=row_idx, column=company_col).value or "").strip().lower()
        company_norm = normalize_company_name(company)
        if company == wanted or (wanted_norm and company_norm == wanted_norm):
            if args.got_ipo and got_col:
                ws.cell(row=row_idx, column=got_col).value = SUPPORTED[args.got_ipo]
            if args.applied and applied_col:
                ws.cell(row=row_idx, column=applied_col).value = SUPPORTED[args.applied]
            updated_rows += 1

    if updated_rows == 0:
        raise RuntimeError(f"Company not found in workbook: {args.company}")

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)

    print(f"Updated {updated_rows} row(s) for company: {args.company}")
    if args.got_ipo:
        print(f"Set got_ipo={SUPPORTED[args.got_ipo] or '(blank)'}")
    if args.applied:
        print(f"Set applied={SUPPORTED[args.applied] or '(blank)'}")
    print(f"Saved workbook: {excel_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
